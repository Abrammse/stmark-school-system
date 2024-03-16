
from django.shortcuts import render,redirect
from django.http import HttpResponseRedirect,HttpResponse,JsonResponse
from student_managment_system.EmailBackEnd import EmailBackEnd
from django.contrib.auth import authenticate
from django.contrib.auth import login,logout
from .models import Events,Staffs,Students,Subjects,Courses,AdminHOD,CustomUser,SessionYearModel,settings,setting2,Attendancestaff,AttendanceReportstaff
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.shortcuts import get_object_or_404
from django.core.files.storage import FileSystemStorage



def index(request):
    all_events = Events.objects.all()
    admin = AdminHOD.objects.all().count()
    staff = Staffs.objects.all().count()
    student=Students.objects.all().count()
    subject=Subjects.objects.all().count()
    setting=settings.object.last()
    settin=setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user.id)


    context = {
        "all_events": all_events,
        "admin": admin,
        "staff": staff,
        "student":student,
        "subject":subject,
        "admin_hod":admin_hod,
        "setting":setting,
        "settin":settin,

    }
    return render(request, 'index.html',context )

def all_events(request):
    all_events = Events.objects.all()
    out = []
    for event in all_events:
        out.append({
            'title': event.name,
            'id': event.id,
            'start': event.start.strftime("%m/%d/%Y, %H:%M:%S"),
            'end': event.end.strftime("%m/%d/%Y, %H:%M:%S"),
        })

    return JsonResponse(out, safe=False)


def add_event(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    title = request.GET.get("title", None)
    event = Events(name=str(title), start=start, end=end)
    event.save()
    data = {}
    return JsonResponse(data)


def update(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    title = request.GET.get("title", None)
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.start = start
    event.end = end
    event.name = title
    event.save()
    data = {}
    return JsonResponse(data)


def remove(request):
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.delete()
    data = {}
    return JsonResponse(data)


def add_staff(request):
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"enter tester.html",{ "admin_hod":admin_hod,  "setting":setting,"settin":settin,})

from django.urls import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import messages
from .models import CustomUser

def add_staff_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        name=request.POST.get("name")
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        username=request.POST.get("username")
        email=request.POST.get("email")
        password=request.POST.get("password")
        address=request.POST.get("address")
        job=request.POST.get("job")
        profile_pic = request.FILES.get("profile_pic")



        try:
            user=CustomUser.objects.create_user(username=username,password=password,email=email,first_name=first_name,user_type=2)
            user.staffs.address=address
            user.staffs.job=job
            user.staffs.name=name
            if profile_pic:
                staff_model.profile_pic = profile_pic
            user.save()
            messages.success(request,"Successfully Added Staff")
            return HttpResponseRedirect(reverse("add_staff"))
        except:
            messages.error(request,"Failed to Add Staff")
            return HttpResponseRedirect(reverse("add_staff"))
def add_admin(request):
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"add_admin.html",{"admin_hod":admin_hod,  "setting":setting,"settin":settin,})
def add_min(request):
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"enter tester2.html",{"admin_hod":admin_hod,  "setting":setting,"settin":settin,})

def add_admin_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        first_name=request.POST.get("first_name")
        name=request.POST.get("name")
        last_name=request.POST.get("last_name")
        username=request.POST.get("username")
        email=request.POST.get("email")
        password=request.POST.get("password")
        address=request.POST.get("address")
        job=request.POST.get("job")
        profile_pic = request.FILES.get("profile_pic")

        try:
            user=CustomUser.objects.create_user(username=username,password=password,email=email,first_name=first_name,user_type=1)
            user.adminhod.address=address
            user.adminhod.job=job
            user.adminhod.name=name
            user.adminhod.profile_pic = profile_pic
            user.save()
            messages.success(request,"Successfully Added admin")
            return HttpResponseRedirect(reverse("add_admin"))
        except:
            messages.error(request,"Failed to Add admin")
            return HttpResponseRedirect(reverse("add_admin"))
def add_student_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        name=request.POST.get("name")
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("bvxvc")
        username=request.POST.get("username")
        email=request.POST.get("email")
        password=request.POST.get("password")
        address=request.POST.get("address")
        job=request.POST.get("job")
        sex=request.POST.get("sex")
        course=request.POST.get("course")
        session_year_id=request.POST.get("session_year_id")
        profile_pic = request.FILES.get("profile_pic")

        try:
            user=CustomUser.objects.create_user(username=username,password=password,email=email,first_name=first_name,user_type=3)
            user.students.address = address
            course_obj = Courses.objects.get(id=course)
            user.students.course_id = course_obj
            user.students.name = name
            user.students.gender = sex
            session_year = SessionYearModel.object.get(id=session_year_id)
            user.students.session_year_id = session_year
            if profile_pic:
                user.students.profile_pic = profile_pic
            user.save()
            messages.success(request,"Successfully Added admin")
            return HttpResponseRedirect(reverse("add_student"))
        except Exception as e:
            messages.error(request,{e})
            print({e})
            return HttpResponseRedirect(reverse("add_student"))
import openpyxl
import openpyxl
from openpyxl.utils import get_column_letter
import os


from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from .models import CustomUser, Courses, SessionYearModel
import openpyxl

from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from .models import CustomUser, Courses, SessionYearModel
import openpyxl
import openpyxl

def add_student_saveexel(request):
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
    else:
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "No file uploaded")
            return HttpResponseRedirect(reverse("add_student"))

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Add error handling to ensure the row has the expected number of columns
                if len(row) < 9:
                    messages.warning(request, f"Skipping row {row}, not enough columns")
                    continue


                first_name = row[1]
                name = row[2]
                course = row[3]
                email = row[4]
                username = row[5]
                gender=row[6]
                password = row[7]
                session_year_id=row[8]
                address=row[9]
                user_exists = CustomUser.objects.filter(username=username).exists()
                if user_exists:
                        existing_user = CustomUser.objects.get(username=username)
                        # Update other columns as needed
                        existing_user.email = email
                        existing_user.first_name = first_name
                        existing_user.save()
                        messages.info(request, f"User '{username}' updated")

                        student = existing_user.students  # Assuming you have a related_name set in the Student model
                        if student:
                            student.name = name
                            user.students.address = address
                            course_obj = Courses.objects.get(id=course)
                            student.course_id = course_obj
                            student.gender = gender
                            session_year = SessionYearModel.object.get(id=session_year_id)
                            student.session_year_id = session_year
                            student.save()

                        # You can choose to update the related Student model here as well

                else:

                          user = CustomUser.objects.create_user(username=username, password=password, email=email, first_name=first_name, user_type=3)
                          user.students.name = name
                          user.students.address = address
                          course_obj = Courses.objects.get(id=course)
                          user.students.course_id = course_obj
                          user.students.gender = gender
                          session_year = SessionYearModel.object.get(id=session_year_id)
                          user.students.session_year_id = session_year
                          user.save()

            messages.success(request, "Students added successfully")
            return HttpResponseRedirect(reverse("add_student"))
        except Exception as e:
            messages.error(request, f"Error adding students: {str(e)}")
            return HttpResponseRedirect(reverse("add_student"))

from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from .models import CustomUser

def add_staff_saveexel(request):
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
    else:
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "No file uploaded")
            return HttpResponseRedirect(reverse("add_student"))

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Add error handling to ensure the row has the expected number of columns
                if len(row) < 7:
                    messages.warning(request, f"Skipping row {row}, not enough columns")
                    continue
                first_name = row[1]
                name = row[2]
                email = row[3]
                username = row[4]
                password = row[5]
                job = row[6]
                address = row[7]



                user_exists = CustomUser.objects.filter(username=username).exists()
                if user_exists:
                    existing_user = CustomUser.objects.get(username=username)
                    existing_user.email = email
                    existing_user.first_name = first_name
                    existing_user.save()
                    messages.info(request, f"User '{username}' updated")

                    staff = existing_user.staffs  # Assuming you have a related_name set in the Staff model
                    if staff:
                        staff.name = name
                        staff.address = address
                        staff.job = job
                        staff.save()

                        # You can choose to update the related Staff model here as well

                else:
                    user = CustomUser.objects.create_user(username=username, password=password, email=email, first_name=first_name, user_type=2)
                    staff = user.staffs
                    if staff:
                        staff.name = name
                        staff.address = address
                        staff.job = job
                        staff.save()

            messages.success(request, "Staff members added/updated successfully")
            return HttpResponseRedirect(reverse("add_staff"))
        except Exception as e:
            messages.error(request, f"Error adding/updating staff members: {str(e)}")
            return HttpResponseRedirect(reverse("add_staff"))

def manage_admin(request):
    staffs=AdminHOD.objects.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"view admin.html",{"staffs":staffs,"admin_hod":admin_hod,  "setting":setting,"settin":settin,})

def add_course(request):
    students=CustomUser.objects.filter(user_type=3)
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"add course.html",{'students': students,"admin_hod":admin_hod,  "setting":setting,"settin":settin})

def add_course_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        course=request.POST.get("course")
        try:
            course_model=Courses(course_name=course)
            course_model.save()
            messages.success(request,"Successfully Added Course")
            return HttpResponseRedirect(reverse("add_course"))
        except:
            messages.error(request,"Failed To Add Course")
            return HttpResponseRedirect(reverse("add_course"))

def add_student(request):
    courses = Courses.objects.all()
    session=SessionYearModel.object.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"add_student.html",{ 'courses': courses,'session':session,"admin_hod":admin_hod, "setting":setting,"settin":settin})



def add_subject(request):
    courses=Courses.objects.all()
    staffs=CustomUser.objects.filter(user_type="2")
    sta=Staffs.objects.all()
    session=SessionYearModel.object.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"add subjects.html",{"staffs":staffs,"courses":courses,"session":session ,"sta":sta,"admin_hod":admin_hod,  "setting":setting,"settin":settin})

def add_subject_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        subject_name=request.POST.get("subject")
        course_id=request.POST.get("course")
        course=Courses.objects.get(id=course_id)
        staff_id=request.POST.get("staff")
        staff=CustomUser.objects.get(id=staff_id)
        session_year_id=request.POST.get("session_year_id")
        session_year_id=SessionYearModel.object.get(id=session_year_id)



        try:
            subject=Subjects(subject_name=subject_name,course_id=course,staff_id=staff,session_year_id=session_year_id)
            subject.save()
            messages.success(request,"Successfully Added Subject")
            return HttpResponseRedirect(reverse("add_subject"))
        except:
            messages.error(request,"Failed to Add Subject")
            return HttpResponseRedirect(reverse("add_subject"))


def manage_staff(request):
    staffs=Staffs.objects.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"view staff.html",{"staffs":staffs,"admin_hod":admin_hod,  "setting":setting,"settin":settin})

def manage_student(request):
    students=Students.objects.all()
    courses = Courses.objects.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"view student.html",{ 'students':students,'courses':courses,"admin_hod":admin_hod,  "setting":setting,"settin":settin})

def manage_course(request):
    courses=Courses.objects.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"view course.html",{"courses":courses,"admin_hod":admin_hod,  "setting":setting,"settin":settin})



def manage_subject(request):
    subjects=Subjects.objects.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"view subject.html",{"subjects":subjects,"admin_hod":admin_hod,  "setting":setting,"settin":settin})


def edit_subject(request,subject_id):
    subject=Subjects.objects.get(id=subject_id)
    courses=Courses.objects.all()
    staffs=CustomUser.objects.filter(user_type=2)
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"hod_template/edit_subject_template.html",{"subject":subject,"staffs":staffs,"courses":courses,"id":subject_id,"admin_hod":admin_hod,  "setting":setting,"settin":settin})


def edit_subject_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        subject_id=request.POST.get("subject_id")
        subject_name=request.POST.get("subject_name")
        staff_id=request.POST.get("staff")
        course_id=request.POST.get("course")


        try:
            subject=Subjects.objects.get(id=subject_id)
            subject.subject_name=subject_name
            staff=CustomUser.objects.get(id=staff_id)
            subject.staff_id=staff
            course=Courses.objects.get(id=course_id)
            subject.course_id=course
            subject.save()

            messages.success(request,"Successfully Edited Subject")
            return HttpResponseRedirect(reverse("edit_subject",kwargs={"subject_id":subject_id}))
        except:
            messages.error(request,"Failed to Edit Subject")
            return HttpResponseRedirect(reverse("edit_subject",kwargs={"subject_id":subject_id}))

def edit_staff(request,staff_id):
    staff=Staffs.objects.get(admin=staff_id)
    setting = settings.object.last()
    settin = setting2.object.last()
    if request.user.is_authenticated:
        try:
            # Attempt to retrieve CustomUser object based on the user's ID
            admin_hod = get_object_or_404(CustomUser, id=request.user.id)
            return render(request, "edit_staff.html",
                          {"staff": staff, "id": staff_id, "admin_hod": admin_hod,
                           "setting": setting, "settin": setting2})
        except CustomUser.DoesNotExist:
            # Handle the case where the CustomUser object does not exist
            pass
    else:

        return render(request, "SingUp.html")
def edit_admin(request,staff_id):
    staff=AdminHOD.objects.get(admin=staff_id)
    setting = settings.object.last()
    settin = setting2.object.last()
    if request.user.is_authenticated:
        try:
            # Attempt to retrieve CustomUser object based on the user's ID
            admin_hod = get_object_or_404(CustomUser, id=request.user.id)
            return render(request, "edit_admin.html",
                          {"staff": staff, "id": staff_id,  "admin_hod": admin_hod,
                           "setting": setting, "settin": setting2})
        except CustomUser.DoesNotExist:
            # Handle the case where the CustomUser object does not exist
            pass
    else:

        return render(request,"SingUp.html")

def edit_student(request,staff_id):
    courses=Courses.objects.all()
    staff=Students.objects.get(admin=staff_id)
    setting = settings.object.last()
    settin = setting2.object.last()
    if request.user.is_authenticated:
        try:
            # Attempt to retrieve CustomUser object based on the user's ID
            admin_hod = get_object_or_404(CustomUser, id=request.user.id)
            return render(request, "edit_student.html",
                          {"staff": staff, "id": staff_id, "courses": courses, "admin_hod": admin_hod,
                           "setting": setting, "settin": settin})
        except CustomUser.DoesNotExist:
            # Handle the case where the CustomUser object does not exist
            pass
    else:
        # Handle non-authenticated users
        # You might redirect the user to the login page or show a different page
        return render(request,"SingUp.html")
def edit_staff_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        staff_id=request.POST.get("staff_id")
        first_name=request.POST.get("first_name")
        email=request.POST.get("email")
        username=request.POST.get("username")
        address=request.POST.get("address")
        job=request.POST.get("job")
        name=request.POST.get("name")
        profile_pic = request.FILES.get("profile_pic")






        try:
            user=CustomUser.objects.get(id=staff_id)
            user.first_name=first_name
            user.email=email
            password = request.POST.get("password")
            if password:
                user.set_password(password)

            user.save()

            staff_model=Staffs.objects.get(admin=staff_id)
            staff_model.address=address
            staff_model.job=job
            staff_model.name=name
            if profile_pic:
                staff_model.profile_pic = profile_pic
            staff_model.save()
            messages.success(request,"Successfully Edited Staff")
            return HttpResponseRedirect(reverse("edit_staff",kwargs={"staff_id":staff_id}))
        except:
            messages.error(request,"Failed to Edit Staff")
            return HttpResponseRedirect(reverse("edit_staff",kwargs={"staff_id":staff_id}))


from django.shortcuts import render, get_object_or_404, reverse, HttpResponseRedirect
from django.contrib import messages
from .models import CustomUser, AdminHOD
# views.py

# views.py

from django.http import HttpResponse
from openpyxl import Workbook
from django.utils import timezone

# views.py

from django.http import HttpResponse
from openpyxl import Workbook
from django.utils import timezone


def export_data_to_excel(request):
    # Fetch the data you want to export (similar to how you're fetching it in your HTML template)
    staffs = AdminHOD.objects.all()  # Replace StaffModel with your actual model name

    # Create a new Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # Define headers for the Excel file
    headers = ['ID', 'Username', 'Profile Picture', 'Name', 'First Name', 'Job', 'Address', 'Email', 'Last Login',
               'Date Joined']
    ws.append(headers)

    # Add data rows to the worksheet
    for staff in staffs:
        last_login = staff.admin.last_login.astimezone(timezone.get_current_timezone()).replace(tzinfo=None) if staff.admin.last_login else None
        date_joined = staff.admin.date_joined.astimezone(timezone.get_current_timezone()).replace(tzinfo=None) if staff.admin.date_joined else None

        row = [
            staff.admin.id,
            staff.admin.username,
            staff.profile_pic.url if staff.profile_pic else 'لايوجد صورة',
            staff.name,
            staff.admin.first_name,
            staff.job,
            staff.address,
            staff.admin.email,
            date_joined
        ]
        ws.append(row)

    # Create the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=staff_data.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


def edit_admin_save(request):
    if request.method != "POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        staff_id = request.POST.get("staff_id")
        first_name = request.POST.get("first_name")
        email = request.POST.get("email")
        username = request.POST.get("username")
        address = request.POST.get("address")
        password = request.POST.get("password")
        job = request.POST.get("job")
        name = request.POST.get("name")
        profile_pic = request.FILES.get("profile_pic")

        try:
            user = CustomUser.objects.get(id=staff_id)
            user.first_name = first_name
            user.email = email
            user.username = username

            staff_model = AdminHOD.objects.get(admin=staff_id)
            staff_model.address = address
            staff_model.job = job
            if profile_pic:
                staff_model.profile_pic = profile_pic
            staff_model.name = name

            if password:
                user.set_password(password)

            user.save()
            staff_model.save()

            messages.success(request, "Successfully Edited Admin")
            return HttpResponseRedirect(reverse("edit_admin", kwargs={"staff_id": staff_id}))

        except CustomUser.DoesNotExist:
            messages.error(request, "Admin user does not exist")
            return HttpResponseRedirect(reverse("edit_admin", kwargs={"staff_id": staff_id}))

        except AdminHOD.DoesNotExist:
            messages.error(request, "AdminHOD object does not exist")
            return HttpResponseRedirect(reverse("edit_admin", kwargs={"staff_id": staff_id}))


def edit_student_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        staff_id=request.POST.get("staff_id")
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        username=request.POST.get("username")
        email=request.POST.get("email")
        password=request.POST.get("password")
        address=request.POST.get("address")
        job=request.POST.get("job")
        sex=request.POST.get("sex")
        name=request.POST.get("name")
        session_start=request.POST.get("session_start")
        session_end=request.POST.get("session_end")
        course=request.POST.get("course")
        profile_pic = request.FILES.get("profile_pic")

        try:
            user = CustomUser.objects.get(id=staff_id)
            user.first_name = first_name
            user.email = email
            user.username = username
            if password:
                user.set_password(password)

            user.save()

            staff_model=Students.objects.get(admin=staff_id)
            staff_model.address = address
            course_obj = Courses.objects.get(id=course)
            staff_model.course_id = course_obj
            staff_model.session_start_year = session_start
            staff_model.session_end_year = session_end
            staff_model.gender = sex
            if profile_pic:
                staff_model.profile_pic = profile_pic
            staff_model.name = name
            staff_model.save()
            messages.success(request,"Successfully Added admin")
            return HttpResponseRedirect(reverse("edit_student" ,kwargs={"staff_id":staff_id}))
        except Exception as e:
            messages.error(request,{e})
            return HttpResponseRedirect(reverse("edit_student",kwargs={"staff_id":staff_id}))



from django.shortcuts import render
from django.shortcuts import render
from .models import AdminHOD

from django.shortcuts import render, get_object_or_404
from django.contrib import messages
from .models import AdminHOD


def edit_course(request,course_id):
    course=Courses.objects.get(id=course_id)
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"hod_template/edit_course_template.html",{"course":course,"id":course_id,"admin_hod":admin_hod,  "setting":setting,"settin":settin})

def edit_course_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        course_id=request.POST.get("course_id")
        course_name=request.POST.get("course")

        try:
            course=Courses.objects.get(id=course_id)
            course.course_name=course_name
            course.save()
            messages.success(request,"Successfully Edited Course")
            return HttpResponseRedirect(reverse("edit_course",kwargs={"course_id":course_id}))
        except:
            messages.error(request,"Failed to Edit Course")
            return HttpResponseRedirect(reverse("edit_course",kwargs={"course_id":course_id}))

from django.shortcuts import get_object_or_404, redirect


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Staffs

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Staffs

def delete_staff(request, pk):
    try:
        staff = get_object_or_404(CustomUser, id=pk)
        staff.delete()
        messages.success(request, "Successfully Deleted Staff")
    except Staffs.DoesNotExist:
        messages.error(request, "Staff not found.")
    return redirect('manage_staff')

def delete_admin(request, pk):
    try:
        staff = get_object_or_404(CustomUser, id=pk)
        staff.delete()
        messages.success(request, "Successfully Deleted Staff")
    except Staffs.DoesNotExist:
        messages.error(request, "Staff not found.")
    return redirect('manage_admin')


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import CustomUser  # Import your CustomUser model

def delete_student(request, pk):
    try:
        student = get_object_or_404(CustomUser, id=pk)
        student.delete()
        messages.success(request, "Successfully deleted student.")
    except CustomUser.DoesNotExist:
        messages.error(request, "Student not found.")
    return redirect('manage_student')

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Courses

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Courses

from django.db.models.deletion import ProtectedError

def delete_course(request, pk):
    try:
        course = Courses.objects.get(pk=pk)
        course.delete()
        messages.success(request, "Successfully deleted course")
    except Courses.DoesNotExist:
        messages.error(request, "Course not found")
    except ProtectedError:
        messages.error(request, "Cannot delete course because related records exist.")
    return redirect('manage_course')

def delete_subject(request, pk):
    try:
        course = Subjects.objects.get(pk=pk)
        course.delete()
        messages.success(request, "Successfully deleted course")
    except Courses.DoesNotExist:
        messages.error(request, "Course not found")
    except ProtectedError:
        messages.error(request, "Cannot delete course because related records exist.")
    return redirect('manage_subject')


def add_session(request):
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)

    return render(request,"add session.html",{"admin_hod":admin_hod,  "setting":setting,"settin":settin})

def manage_session(request):
    session=SessionYearModel.object.all()
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"view session.html",{"session":session,"admin_hod":admin_hod,  "setting":setting,"settin":settin})

def chrange_settings(request):
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"chrange settings.html",{"admin_hod":admin_hod,  "setting":setting,"settin":settin})
def add_session_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("manage_session"))
    else:
        session_start_year=request.POST.get("session_start")
        session_end_year=request.POST.get("session_end")
        Term=request.POST.get("Term")


        try:
            sessionyear=SessionYearModel(session_start_year=session_start_year,session_end_year=session_end_year,Term=Term)
            sessionyear.save()
            messages.success(request, "Successfully Added Session")
            return HttpResponseRedirect(reverse("add_session"))
        except:
            messages.error(request, "Failed to Add Session")
            return HttpResponseRedirect(reverse("add_session"))
def chrange_settings_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("manage_session"))
    else:
        title=request.POST.get("title")



        try:
            sessionyear=settings(title=title)
            sessionyear.save()
            messages.success(request, "Successfully Added Session")
            return HttpResponseRedirect(reverse("chrange_settings"))
        except:
            messages.error(request, "Failed to Add Session")
            return HttpResponseRedirect(reverse("chrange_settings"))

def chrange_settings2(request):
    setting = settings.object.last()
    settin = setting2.object.last()
    admin_hod = AdminHOD.objects.get(admin=request.user)
    return render(request,"chrange settings title.html",{"admin_hod":admin_hod,  "setting":setting,"settin":settin})
def chrange_setting2_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("manage_session"))

    else:
        profile_pic = request.FILES['profile_pic']
        fs = FileSystemStorage()
        filename = fs.save(profile_pic.name, profile_pic)
        profile_pic_url = fs.url(filename)

        try:
            sessionyear = setting2.object.create(pic=profile_pic_url)
            sessionyear.save()
            messages.success(request, "Successfully Added Session")
            return HttpResponseRedirect(reverse("chrange_settings2"))
        except:
            messages.error(request, "Failed to Add Session")
            return HttpResponseRedirect(reverse("chrange_settings2"))
def staff_take_attendance_staff(request):
    setting=settings.object.last()
    staff_id=AdminHOD.objects.get(admin=request.user.id)
    session_years = SessionYearModel.object.all()
    action=request.GET.get('action')
    session_obj=None
    staffsr=None
    if action is  not None:
        if request.method== "POST":
            session=request.POST.get("session")

            session_obj=SessionYearModel.object.get(id=session)
            staffsr=Staffs.objects.all()
    return render(request,"add Attandance-staff.html",{"session_years":session_years , "session_obj":session_obj,"action":action,"staffsr":staffsr,"setting":setting,})



def staff_take_attendance_save_staff(request):
    if request.method == "POST":
        session = request.POST.get("session")
        attendance_date = request.POST.get("attendance_date")
        selected_student_ids = request.POST.getlist("student_id")  # Retrieve list of selected student IDs

        # Retrieve subject and session objects
        get_session_year = SessionYearModel.object.get(id=session)

        # Check if attendance for the given subject, session, and date already exists
        attendance, created = Attendancestaff.objects.get_or_create(
            attendance_date=attendance_date,
            session_year_id=get_session_year
        )

        # Get existing attendance reports for the given attendance
        existing_reports = AttendanceReportstaff.objects.filter(attendance_id=attendance)

        # Create a dictionary to store existing reports by student ID
        existing_reports_dict = {report.student_id.id: report for report in existing_reports}

        # Iterate through selected student IDs
        for student_id in selected_student_ids:
            student = Staffs.objects.get(id=student_id)

            # Check if the student has an existing report
            if student.id in existing_reports_dict:
                # If the student already has a report, update its status
                existing_report = existing_reports_dict[student.id]
                existing_report.status = True
                existing_report.save()
            else:
                # If the student doesn't have a report, create a new one with status=True
                AttendanceReportstaff.objects.create(
                    student_id=student,
                    attendance_id=attendance,
                    status=True,
                )

        messages.success(request, "Attendance successfully recorded")
        return HttpResponseRedirect(reverse("staff_take_attendance_staff"))
    else:
        # Handle GET request
        pass

from django.shortcuts import render
from .models import SessionYearModel, Attendancestaff, AttendanceReportstaff, CustomUser, Subjects

def staff_view_attendance_staff(request):
    session_years = SessionYearModel.object.all()
    attendance = Attendancestaff.objects.filter(session_year_id__in=session_years)

    action = request.GET.get('action')
    session_obj = None
    attendance_obj = None
    attendance_report = None
    staff_not_found = None

    if action is not None:
        if request.method == "POST":
            session_id = request.POST.get("session")
            attendance_id = request.POST.get("attendance")

            session_obj = SessionYearModel.object.get(id=session_id)
            attendance_obj = Attendancestaff.objects.get(id=attendance_id)

            attendance_report = AttendanceReportstaff.objects.filter(attendance_id=attendance_id)

            # Get the list of staff members who are registered to take attendance
            staff_registered = Staffs.objects.all()

            # Get the list of staff members who attended the class
            staff_attended_ids = [report.student_id.id for report in attendance_report]

            # Find staff members who are registered but not found in the attendance report
            staff_not_found = staff_registered.exclude(id__in=staff_attended_ids)

    return render(request, "view Attandance-staff.html", {
        "session_years": session_years,
        "session_obj": session_obj,
        "attendance": attendance,
        "action": action,
        "attendance_obj": attendance_obj,
        "attendance_report": attendance_report,
        "staff_not_found": staff_not_found
    })


